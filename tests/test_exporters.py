"""Tests for Excel and PDF export (exporters.py)."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from expenses_tracker.exporters import export_reports
from tests.conftest import CATEGORY_ROWS, MONTH_ROWS, TRANSACTIONS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _run(tmp_path: Path, fmt: str = "all", language: str = "en") -> list[Path]:
    return export_reports(
        transactions=TRANSACTIONS,
        category_rows=CATEGORY_ROWS,
        month_rows=MONTH_ROWS,
        output_dir=tmp_path,
        fmt=fmt,
        language=language,
    )


# ---------------------------------------------------------------------------
# Output directory
# ---------------------------------------------------------------------------


class TestOutputDirectory:
    def test_creates_directory_if_missing(self, tmp_path):
        out = tmp_path / "reports_subdir"
        assert not out.exists()
        export_reports(TRANSACTIONS, CATEGORY_ROWS, MONTH_ROWS, output_dir=out, fmt="excel")
        assert out.is_dir()


# ---------------------------------------------------------------------------
# Return value
# ---------------------------------------------------------------------------


class TestReturnValue:
    def test_returns_list_of_paths(self, tmp_path):
        result = _run(tmp_path)
        assert isinstance(result, list)
        assert all(isinstance(p, Path) for p in result)

    def test_all_format_returns_six_files(self, tmp_path):
        result = _run(tmp_path, fmt="all")
        assert len(result) == 6

    def test_excel_only_returns_one_file(self, tmp_path):
        result = _run(tmp_path, fmt="excel")
        assert len(result) == 1

    def test_csv_only_returns_one_file(self, tmp_path):
        result = _run(tmp_path, fmt="csv")
        assert len(result) == 1

    def test_pdf_only_returns_one_file(self, tmp_path):
        result = _run(tmp_path, fmt="pdf")
        assert len(result) == 1

    def test_returned_paths_exist(self, tmp_path):
        for path in _run(tmp_path):
            assert path.exists()

    def test_empty_transactions_still_generates_files(self, tmp_path):
        # The empty-data guard lives in cli.py, not here; exporters always produces output.
        result = export_reports([], [], [], output_dir=tmp_path, fmt="all")
        assert len(result) == 6


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


class TestExcelExport:
    def test_creates_xlsx_file(self, tmp_path):
        (path,) = _run(tmp_path, fmt="excel")
        assert path.suffix == ".xlsx"

    def test_excel_file_is_non_empty(self, tmp_path):
        (path,) = _run(tmp_path, fmt="excel")
        assert path.stat().st_size > 0

    def test_excel_has_transactions_sheet(self, tmp_path):
        import openpyxl
        (path,) = _run(tmp_path, fmt="excel")
        wb = openpyxl.load_workbook(path)
        assert "Transactions" in wb.sheetnames

    def test_excel_has_by_category_sheet(self, tmp_path):
        import openpyxl
        (path,) = _run(tmp_path, fmt="excel")
        wb = openpyxl.load_workbook(path)
        assert "ByCategory" in wb.sheetnames

    def test_excel_has_by_month_sheet(self, tmp_path):
        import openpyxl
        (path,) = _run(tmp_path, fmt="excel")
        wb = openpyxl.load_workbook(path)
        assert "ByMonth" in wb.sheetnames

    def test_transactions_sheet_row_count(self, tmp_path):
        import openpyxl
        (path,) = _run(tmp_path, fmt="excel")
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        # 1 header row + len(TRANSACTIONS) data rows
        assert ws.max_row == 1 + len(TRANSACTIONS)

    def test_spanish_excel_does_not_raise(self, tmp_path):
        result = _run(tmp_path, fmt="excel", language="es")
        assert len(result) == 1

    def test_excel_escapes_formula_like_user_text(self, tmp_path):
        import openpyxl

        transactions = [
            {
                "id": 1,
                "amount": 10.0,
                "transaction_type": "expense",
                "category": "=cmd|' /C calc'!A0",
                "transaction_date": "2026-04-29",
                "description": "+SUM(1,1)",
                "created_at": "@malicious",
            }
        ]

        (path,) = export_reports(transactions, [], [], output_dir=tmp_path, fmt="excel")
        workbook = openpyxl.load_workbook(path, data_only=False)
        row = next(workbook["Transactions"].iter_rows(min_row=2, max_row=2, values_only=True))

        assert row[3].startswith("'=")
        assert row[5].startswith("'+")
        assert row[6].startswith("'@")


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


class TestCsvExport:
    def test_creates_csv_file(self, tmp_path):
        (path,) = _run(tmp_path, fmt="csv")
        assert path.suffix == ".csv"

    def test_csv_file_is_non_empty(self, tmp_path):
        (path,) = _run(tmp_path, fmt="csv")
        assert path.stat().st_size > 0

    def test_csv_contains_transactions(self, tmp_path):
        (path,) = _run(tmp_path, fmt="csv")
        content = path.read_text(encoding="utf-8-sig")
        assert "Salario" in content
        assert "Freelance" in content

    def test_csv_escapes_formula_like_user_text(self, tmp_path):
        transactions = [
            {
                "id": 1,
                "amount": 10.0,
                "transaction_type": "expense",
                "category": "=cmd|' /C calc'!A0",
                "transaction_date": "2026-04-29",
                "description": "+SUM(1,1)",
                "created_at": "@malicious",
            }
        ]

        (path,) = export_reports(transactions, [], [], output_dir=tmp_path, fmt="csv")
        content = path.read_text(encoding="utf-8-sig")
        assert "'=cmd" in content
        assert "'+SUM" in content
        assert "'@malicious" in content


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------


class TestPdfExport:
    def test_creates_pdf_file(self, tmp_path):
        (path,) = _run(tmp_path, fmt="pdf")
        assert path.suffix == ".pdf"

    def test_pdf_file_is_non_empty(self, tmp_path):
        (path,) = _run(tmp_path, fmt="pdf")
        assert path.stat().st_size > 0

    def test_pdf_starts_with_pdf_magic_bytes(self, tmp_path):
        (path,) = _run(tmp_path, fmt="pdf")
        assert path.read_bytes()[:4] == b"%PDF"

    def test_spanish_pdf_does_not_raise(self, tmp_path):
        result = _run(tmp_path, fmt="pdf", language="es")
        assert len(result) == 1


# ---------------------------------------------------------------------------
# JSON export
# ---------------------------------------------------------------------------


class TestJsonExport:
    def test_creates_json_file(self, tmp_path):
        (path,) = _run(tmp_path, fmt="json")
        assert path.suffix == ".json"

    def test_json_contains_transactions(self, tmp_path):
        (path,) = _run(tmp_path, fmt="json")
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "transactions" in data
        assert len(data["transactions"]) == len(TRANSACTIONS)

    def test_json_has_meta(self, tmp_path):
        (path,) = _run(tmp_path, fmt="json")
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data["meta"]["language"] == "en"


# ---------------------------------------------------------------------------
# YAML export
# ---------------------------------------------------------------------------


class TestYamlExport:
    def test_creates_yaml_file(self, tmp_path):
        (path,) = _run(tmp_path, fmt="yaml")
        assert path.suffix == ".yaml"

    def test_yaml_contains_transactions(self, tmp_path):
        import yaml

        (path,) = _run(tmp_path, fmt="yaml")
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        assert "transactions" in data
        assert len(data["transactions"]) == len(TRANSACTIONS)


# ---------------------------------------------------------------------------
# HTML export
# ---------------------------------------------------------------------------


class TestHtmlExport:
    def test_creates_html_file(self, tmp_path):
        (path,) = _run(tmp_path, fmt="html")
        assert path.suffix == ".html"

    def test_html_is_non_empty(self, tmp_path):
        (path,) = _run(tmp_path, fmt="html")
        assert path.stat().st_size > 0

    def test_html_contains_plotly(self, tmp_path):
        (path,) = _run(tmp_path, fmt="html")
        content = path.read_text(encoding="utf-8")
        assert "plotly" in content.lower()


# ---------------------------------------------------------------------------
# Monthly PDF export
# ---------------------------------------------------------------------------


class TestMonthlyPdfExport:
    def test_creates_pdf_file(self, tmp_path):
        (path,) = export_reports(
            TRANSACTIONS, CATEGORY_ROWS, MONTH_ROWS, output_dir=tmp_path, fmt="monthly_pdf"
        )
        assert path.suffix == ".pdf"
        assert path.read_bytes()[:4] == b"%PDF"

    def test_uses_provided_year_month(self, tmp_path):
        (path,) = export_reports(
            TRANSACTIONS, CATEGORY_ROWS, MONTH_ROWS, output_dir=tmp_path, fmt="monthly_pdf", year_month="2025-01"
        )
        assert path.name.startswith("monthly_report_2025-01")

    def test_empty_transactions_raises(self, tmp_path):
        with pytest.raises(ValueError):
            export_reports([], [], [], output_dir=tmp_path, fmt="monthly_pdf")
