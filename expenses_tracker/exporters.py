from __future__ import annotations

import csv
import io
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape


def _escape_reportlab(value: str) -> str:
    """Escape text for ReportLab Paragraph content, stripping HTML tags."""
    import re
    text = re.sub(r"<[^>]+>", "", value)
    text = escape(text)
    return text

import yaml
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from expenses_tracker.i18n import tr
from expenses_tracker.security import apply_private_permissions, sanitize_spreadsheet_text

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    _PLOTLY_AVAILABLE = True
except Exception:  # pragma: no cover
    _PLOTLY_AVAILABLE = False


def export_reports(
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    output_dir: str | Path = "reports",
    fmt: str = "all",
    language: str = "en",
    year_month: str | None = None,
    budget_rows: list[dict[str, Any]] | None = None,
) -> list[Path]:
    """Generate reports in the specified format and return output file paths."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    generated_files: list[Path] = []

    if fmt in {"excel", "all"}:
        excel_file = output_path / f"report_{timestamp}.xlsx"
        _export_excel(excel_file, transactions, category_rows, month_rows, language)
        apply_private_permissions(excel_file)
        generated_files.append(excel_file)

    if fmt in {"csv", "all"}:
        csv_file = output_path / f"report_{timestamp}.csv"
        _export_csv(csv_file, transactions, language)
        apply_private_permissions(csv_file)
        generated_files.append(csv_file)

    if fmt in {"pdf", "all"}:
        pdf_file = output_path / f"report_{timestamp}.pdf"
        _export_pdf(pdf_file, transactions, category_rows, month_rows, language, budget_rows)
        apply_private_permissions(pdf_file)
        generated_files.append(pdf_file)

    if fmt in {"json", "all"}:
        json_file = output_path / f"report_{timestamp}.json"
        _export_json(json_file, transactions, category_rows, month_rows, language)
        apply_private_permissions(json_file)
        generated_files.append(json_file)

    if fmt in {"yaml", "all"}:
        yaml_file = output_path / f"report_{timestamp}.yaml"
        _export_yaml(yaml_file, transactions, category_rows, month_rows, language)
        apply_private_permissions(yaml_file)
        generated_files.append(yaml_file)

    if fmt in {"html", "all"}:
        html_file = output_path / f"report_{timestamp}.html"
        _export_html(html_file, transactions, category_rows, month_rows, language)
        apply_private_permissions(html_file)
        generated_files.append(html_file)

    if fmt == "monthly_pdf":
        monthly_pdf_file = output_path / f"monthly_report_{year_month or timestamp}.pdf"
        _export_monthly_pdf(monthly_pdf_file, transactions, language, year_month, budget_rows)
        apply_private_permissions(monthly_pdf_file)
        generated_files.append(monthly_pdf_file)

    return generated_files


def _figure_to_image(figure: Any, width: int = 500, dpi: int = 150) -> RLImage:
    """Convert a matplotlib Figure to a ReportLab Image without disk I/O."""
    buf = io.BytesIO()
    figure.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    img = RLImage(buf, width=width, height=width * 0.5)
    return img


def _compute_savings_rate(income: float, expense: float) -> float:
    """Calculate savings rate as percentage."""
    if income <= 0:
        return 0.0
    return ((income - expense) / income) * 100


def _compute_trend_direction(values: list[float]) -> str:
    """Calculate trend direction: growth, decline, or stable."""
    if len(values) < 2:
        return "stable"
    first_half = sum(values[: len(values) // 2])
    second_half = sum(values[len(values) // 2 :])
    diff_pct = ((second_half - first_half) / first_half * 100) if first_half > 0 else 0
    if diff_pct > 5:
        return "growth"
    if diff_pct < -5:
        return "decline"
    return "stable"


def _detect_anomalies(month_rows: list[dict[str, Any]]) -> list[str]:
    """Detect anomalies: deficit months, spending spikes."""
    anomalies: list[str] = []
    prev_expense: float | None = None
    deficit_count = 0

    for row in month_rows:
        income = float(row.get("income", 0))
        expense = float(row.get("expense", 0))
        if expense > income:
            deficit_count += 1
        if prev_expense is not None and prev_expense > 0:
            pct_change = ((expense - prev_expense) / prev_expense) * 100
            if pct_change > 30:
                anomalies.append(f"spike|{row.get('month', '')}|{pct_change:.0f}")
        prev_expense = expense

    if deficit_count > 0:
        anomalies.insert(0, f"deficit|{deficit_count}")

    surplus_count = len(month_rows) - deficit_count
    if surplus_count > 0:
        anomalies.append(f"surplus|{surplus_count}|{len(month_rows)}")

    return anomalies


def _localize_transaction_type(value: str, language: str) -> str:
    if value == "income":
        return tr(language, "type_income")
    if value == "expense":
        return tr(language, "type_expense")
    return value


def _export_excel(
    output_file: Path,
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    language: str,
) -> None:
    workbook = Workbook()

    movement_sheet = workbook.active
    movement_sheet.title = tr(language, "excel_sheet_transactions")
    movement_sheet.append(
        [
            tr(language, "excel_col_id"),
            tr(language, "excel_col_date"),
            tr(language, "excel_col_type"),
            tr(language, "excel_col_category"),
            tr(language, "excel_col_amount"),
            tr(language, "excel_col_description"),
            tr(language, "excel_col_created"),
        ]
    )
    for row in transactions:
        transaction_type = _localize_transaction_type(str(row["transaction_type"]), language)
        movement_sheet.append(
            [
                row["id"],
                sanitize_spreadsheet_text(row["transaction_date"]),
                sanitize_spreadsheet_text(transaction_type),
                sanitize_spreadsheet_text(row["category"]),
                float(row["amount"] or 0),
                sanitize_spreadsheet_text(row["description"]),
                sanitize_spreadsheet_text(row["created_at"]),
            ]
        )

    category_sheet = workbook.create_sheet(tr(language, "excel_sheet_by_category"))
    category_sheet.append(
        [
            tr(language, "excel_col_category"),
            tr(language, "pdf_col_income"),
            tr(language, "pdf_col_expense"),
            tr(language, "pdf_col_balance"),
        ]
    )
    for row in category_rows:
        category_sheet.append(
            [
                sanitize_spreadsheet_text(row["category"]),
                float(row["income"] or 0),
                float(row["expense"] or 0),
                float(row["balance"] or 0),
            ]
        )

    month_sheet = workbook.create_sheet(tr(language, "excel_sheet_by_month"))
    month_sheet.append(
        [
            tr(language, "chart_x_month"),
            tr(language, "pdf_col_income"),
            tr(language, "pdf_col_expense"),
            tr(language, "pdf_col_balance"),
        ]
    )
    for row in month_rows:
        month_sheet.append(
            [
                sanitize_spreadsheet_text(row["month"]),
                float(row["income"] or 0),
                float(row["expense"] or 0),
                float(row["balance"] or 0),
            ]
        )

    workbook.save(output_file)


def _export_csv(output_file: Path, transactions: list[dict[str, Any]], language: str) -> None:
    with output_file.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(
            [
                tr(language, "excel_col_id"),
                tr(language, "excel_col_date"),
                tr(language, "excel_col_type"),
                tr(language, "excel_col_category"),
                tr(language, "excel_col_amount"),
                tr(language, "excel_col_description"),
                tr(language, "excel_col_created"),
            ]
        )

        for row in transactions:
            transaction_type = _localize_transaction_type(str(row["transaction_type"]), language)
            writer.writerow(
                [
                    row["id"],
                    sanitize_spreadsheet_text(row["transaction_date"]),
                    sanitize_spreadsheet_text(transaction_type),
                    sanitize_spreadsheet_text(row["category"]),
                    f"{float(row['amount'] or 0):.2f}",
                    sanitize_spreadsheet_text(row["description"]),
                    sanitize_spreadsheet_text(row["created_at"]),
                ]
            )


def _export_pdf(
    output_file: Path,
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    language: str,
    budget_rows: list[dict[str, Any]] | None = None,
) -> None:
    from expenses_tracker.charts import generate_all_figures

    base_styles = getSampleStyleSheet()
    styles = _build_pdf_styles(base_styles)
    document = SimpleDocTemplate(
        str(output_file),
        pagesize=A4,
        leftMargin=28,
        rightMargin=28,
        topMargin=36,
        bottomMargin=26,
    )
    elements: list[Any] = []

    summary_metrics = _compute_summary_metrics(transactions, category_rows, month_rows)

    # 1. Premium cover with KPI dashboard
    elements.extend(_build_premium_cover(summary_metrics, styles, language))
    elements.append(PageBreak())

    # 2. Trend analysis with anomalies and projection
    elements.extend(_build_trend_analysis_section(summary_metrics, month_rows, styles, language))
    elements.append(PageBreak())

    # 3. Generate all charts as figures
    figures = generate_all_figures(category_rows, month_rows, language, "default")

    # 4. Embed charts
    chart_title_map = {
        "line": "pdf_chart_evolution",
        "bar": "pdf_chart_categories",
        "pie": "pdf_chart_distribution",
        "scatter": "pdf_chart_scatter",
        "bar3d": "pdf_chart_3d",
        "forecast": "pdf_chart_forecast",
        "sankey": "pdf_chart_sankey",
    }

    for chart_key in ["line", "bar", "pie", "scatter", "bar3d", "forecast", "sankey"]:
        if chart_key in figures:
            elements.extend(_build_chart_section(
                figures[chart_key],
                chart_title_map[chart_key],
                styles,
                language,
            ))

    # 5. Budget section (if data provided)
    if budget_rows:
        elements.extend(_build_budget_section(budget_rows, styles, language))

    # 6. Category table
    category_table_rows = [
        [
            _escape_reportlab(str(row["category"])),
            f"{float(row['income'] or 0):.2f}",
            f"{float(row['expense'] or 0):.2f}",
            f"{float(row['balance'] or 0):.2f}",
        ]
        for row in category_rows
    ]
    _append_paginated_table(
        elements=elements,
        styles=styles,
        title=tr(language, "pdf_table_category_title"),
        header=[
            tr(language, "excel_col_category"),
            tr(language, "pdf_col_income"),
            tr(language, "pdf_col_expense"),
            tr(language, "pdf_col_balance"),
        ],
        rows=category_table_rows,
        chunk_size=28,
        col_widths=[180, 95, 95, 95],
        text_column_indexes=[0],
        language=language,
    )

    # 7. Month table
    month_table_rows = [
        [
            _escape_reportlab(str(row["month"])),
            f"{float(row['income'] or 0):.2f}",
            f"{float(row['expense'] or 0):.2f}",
            f"{float(row['balance'] or 0):.2f}",
        ]
        for row in month_rows
    ]
    _append_paginated_table(
        elements=elements,
        styles=styles,
        title=tr(language, "pdf_table_month_title"),
        header=[
            tr(language, "chart_x_month"),
            tr(language, "pdf_col_income"),
            tr(language, "pdf_col_expense"),
            tr(language, "pdf_col_balance"),
        ],
        rows=month_table_rows,
        chunk_size=30,
        col_widths=[140, 105, 105, 105],
        text_column_indexes=[0],
        language=language,
    )

    # 8. Transactions table
    movement_table_rows = [
        [
            _escape_reportlab(str(row["transaction_date"])),
            _escape_reportlab(_localize_transaction_type(str(row["transaction_type"]), language)),
            _escape_reportlab(str(row["category"])),
            f"{float(row['amount'] or 0):.2f}",
            _escape_reportlab(str(row["description"])),
        ]
        for row in transactions
    ]
    _append_paginated_table(
        elements=elements,
        styles=styles,
        title=tr(language, "pdf_table_transactions_title"),
        header=[
            tr(language, "excel_col_date"),
            tr(language, "excel_col_type"),
            tr(language, "excel_col_category"),
            tr(language, "excel_col_amount"),
            tr(language, "excel_col_description"),
        ],
        rows=movement_table_rows,
        chunk_size=22,
        col_widths=[86, 76, 102, 70, 136],
        text_column_indexes=[0, 1, 2, 4],
        language=language,
    )

    document.build(
        elements,
        onFirstPage=lambda canvas, doc: _draw_page_number(canvas, doc, language),
        onLaterPages=lambda canvas, doc: _draw_page_number(canvas, doc, language),
    )


def _append_paginated_table(
    elements: list[Any],
    styles: Any,
    title: str,
    header: list[str],
    rows: list[list[str]],
    chunk_size: int,
    col_widths: list[int],
    text_column_indexes: list[int],
    language: str,
) -> None:
    if not rows:
        elements.append(Paragraph(title, styles["section_title"]))
        elements.append(Paragraph(tr(language, "pdf_no_data_block"), styles["body_text"]))
        elements.append(Spacer(1, 10))
        return

    for index, chunk in enumerate(_chunks(rows, chunk_size)):
        if index > 0:
            elements.append(PageBreak())

        table_title = title if index == 0 else f"{title} {tr(language, 'pdf_continued', page=index + 1)}"
        elements.append(Paragraph(table_title, styles["section_title"]))
        elements.append(
            _styled_table(
                [header, *chunk],
                col_widths=col_widths,
                text_column_indexes=text_column_indexes,
            )
        )
        elements.append(Spacer(1, 10))


def _styled_table(
    data: list[list[str]],
    col_widths: list[int],
    text_column_indexes: list[int],
) -> Table:
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_commands: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f5f8fc"), colors.white]),
        ("ALIGN", (0, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#b8c4d6")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]

    for column_index in text_column_indexes:
        style_commands.append(("ALIGN", (column_index, 1), (column_index, -1), "LEFT"))

    table.setStyle(TableStyle(style_commands))
    return table


def _compute_summary_metrics(
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    total_income = 0.0
    total_expense = 0.0
    categories: set[str] = set()
    period_start = "N/A"
    period_end = "N/A"

    if transactions:
        dates = [str(row["transaction_date"]) for row in transactions]
        period_start = min(dates)
        period_end = max(dates)

        for row in transactions:
            amount = float(row["amount"] or 0)
            categories.add(str(row["category"]))
            if row["transaction_type"] == "income":
                total_income += amount
            elif row["transaction_type"] == "expense":
                total_expense += amount

    balance = total_income - total_expense
    top_expense = _top_category(category_rows, "expense")
    top_income = _top_category(category_rows, "income")

    return {
        "period_start": period_start,
        "period_end": period_end,
        "transactions_count": len(transactions),
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": balance,
        "categories_count": len(categories),
        "months_count": len(month_rows),
        "top_expense": top_expense,
        "top_income": top_income,
    }


def _build_executive_summary(metrics: dict[str, Any], language: str) -> str:
    lines = [
        tr(
            language,
            "pdf_period_analyzed",
            start=_escape_reportlab(str(metrics["period_start"])),
            end=_escape_reportlab(str(metrics["period_end"])),
        ),
        tr(language, "pdf_transactions_count", count=metrics["transactions_count"]),
        tr(language, "pdf_income_total", amount=float(metrics["total_income"])),
        tr(language, "pdf_expense_total", amount=float(metrics["total_expense"])),
        tr(language, "pdf_balance_total", amount=float(metrics["balance"])),
        tr(language, "pdf_categories_count", count=metrics["categories_count"]),
        tr(language, "pdf_months_count", count=metrics["months_count"]),
        tr(language, "pdf_top_expense", value=_escape_reportlab(str(metrics["top_expense"]))),
        tr(language, "pdf_top_income", value=_escape_reportlab(str(metrics["top_income"]))),
    ]
    return "<br/>".join(lines)


def _build_trend_analysis_section(
    metrics: dict[str, Any],
    month_rows: list[dict[str, Any]],
    styles: dict[str, ParagraphStyle],
    language: str,
) -> list[Any]:
    """Build a trend analysis section with narrative text and anomaly alerts."""
    elements: list[Any] = []
    elements.append(Paragraph(tr(language, "pdf_exec_summary"), styles["section_title"]))

    elements.append(Paragraph(_build_executive_summary(metrics, language), styles["body_text"]))
    elements.append(Spacer(1, 8))

    if month_rows:
        balances = [float(row.get("balance", 0)) for row in month_rows]
        trend = _compute_trend_direction(balances)
        trend_key = f"pdf_trend_{trend}"
        trend_style = f"trend_{trend}"
        elements.append(Paragraph(tr(language, trend_key), styles[trend_style]))
        elements.append(Spacer(1, 6))

        anomalies = _detect_anomalies(month_rows)
        for anomaly in anomalies:
            parts = anomaly.split("|")
            if parts[0] == "deficit":
                elements.append(
                    Paragraph(
                        tr(language, "pdf_alert_deficit", count=int(parts[1])),
                        styles["insight_text"],
                    )
                )
            elif parts[0] == "spike":
                elements.append(
                    Paragraph(
                        tr(language, "pdf_alert_spike", category=parts[1], percent=float(parts[2])),
                        styles["insight_text"],
                    )
                )
            elif parts[0] == "surplus":
                elements.append(
                    Paragraph(
                        tr(language, "pdf_alert_surplus", count=int(parts[1]), total=int(parts[2])),
                        styles["insight_text"],
                    )
                )

        if len(balances) >= 2:
            import numpy as np

            x = np.arange(len(balances))
            coeffs = np.polyfit(x, balances, 1)
            poly = np.poly1d(coeffs)
            projection = float(poly(len(balances)))
            elements.append(Spacer(1, 4))
            elements.append(
                Paragraph(
                    tr(language, "pdf_projection_next", amount=projection),
                    styles["analysis_heading"],
                )
            )

    elements.append(Spacer(1, 10))
    return elements


def _build_chart_section(
    figure: Any,
    title_key: str,
    styles: dict[str, ParagraphStyle],
    language: str,
    width: int = 480,
) -> list[Any]:
    """Build a chart section with title and embedded matplotlib figure."""
    elements: list[Any] = []
    elements.append(PageBreak())
    elements.append(Paragraph(tr(language, title_key), styles["chart_title"]))
    elements.append(Spacer(1, 6))
    elements.append(_figure_to_image(figure, width=width))
    elements.append(Spacer(1, 10))
    return elements


def _build_budget_section(
    budget_rows: list[dict[str, Any]],
    styles: dict[str, ParagraphStyle],
    language: str,
) -> list[Any]:
    """Build a budget analysis section with chart and table."""
    elements: list[Any] = []
    elements.append(PageBreak())
    elements.append(Paragraph(tr(language, "pdf_budget_section_title"), styles["section_title"]))
    elements.append(Spacer(1, 6))

    from expenses_tracker.charts import budget_comparison_figure, get_palette

    colors_palette = get_palette("default")
    figure = budget_comparison_figure(budget_rows, language, colors_palette)
    elements.append(_figure_to_image(figure, width=480))
    elements.append(Spacer(1, 10))

    header = [
        tr(language, "excel_col_category"),
        tr(language, "pdf_budget_planned"),
        tr(language, "pdf_budget_actual"),
        tr(language, "pdf_budget_delta"),
        tr(language, "pdf_budget_compliance"),
    ]

    rows = []
    for row in budget_rows:
        planned = float(row.get("planned", 0))
        actual = float(row.get("actual", 0))
        delta = planned - actual
        compliance = ((planned - actual) / planned * 100) if planned > 0 else 0
        status = tr(language, "pdf_budget_ok") if delta >= 0 else tr(language, "pdf_budget_exceeded")
        rows.append([
            _escape_reportlab(str(row.get("category", ""))),
            f"{planned:.2f}",
            f"{actual:.2f}",
            f"{delta:.2f}",
            f"{compliance:.0f}% - {status}",
        ])

    table_data = [header, *rows]
    table = Table(table_data, colWidths=[120, 80, 80, 80, 120], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f5f8fc"), colors.white]),
        ("ALIGN", (0, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#b8c4d6")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 10))
    return elements


def _build_pdf_styles(base_styles: Any) -> dict[str, ParagraphStyle]:
    return {
        "cover_title": ParagraphStyle(
            "cover_title",
            parent=base_styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=28,
            textColor=colors.white,
        ),
        "cover_subtitle": ParagraphStyle(
            "cover_subtitle",
            parent=base_styles["Normal"],
            fontName="Helvetica",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#e6edf7"),
        ),
        "section_title": ParagraphStyle(
            "section_title",
            parent=base_styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#1f3a5f"),
            spaceAfter=6,
        ),
        "body_text": ParagraphStyle(
            "body_text",
            parent=base_styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#243447"),
        ),
        "kpi_cell": ParagraphStyle(
            "kpi_cell",
            parent=base_styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#0f233d"),
        ),
        "kpi_value": ParagraphStyle(
            "kpi_value",
            parent=base_styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1f3a5f"),
            alignment=1,
        ),
        "kpi_label": ParagraphStyle(
            "kpi_label",
            parent=base_styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#556273"),
            alignment=1,
        ),
        "insight_text": ParagraphStyle(
            "insight_text",
            parent=base_styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#243447"),
            leftIndent=10,
            borderPadding=4,
        ),
        "chart_title": ParagraphStyle(
            "chart_title",
            parent=base_styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#1f3a5f"),
            spaceAfter=4,
        ),
        "trend_growth": ParagraphStyle(
            "trend_growth",
            parent=base_styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#2e7d32"),
        ),
        "trend_decline": ParagraphStyle(
            "trend_decline",
            parent=base_styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#c62828"),
        ),
        "trend_stable": ParagraphStyle(
            "trend_stable",
            parent=base_styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#1565c0"),
        ),
        "header_text": ParagraphStyle(
            "header_text",
            parent=base_styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#8899aa"),
        ),
        "analysis_heading": ParagraphStyle(
            "analysis_heading",
            parent=base_styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            textColor=colors.HexColor("#1f3a5f"),
            spaceBefore=8,
            spaceAfter=4,
        ),
    }


def _build_cover_banner(
    metrics: dict[str, Any],
    styles: dict[str, ParagraphStyle],
    language: str,
) -> Table:
    subtitle = (
        tr(language, "pdf_generated", timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        + "<br/>"
        + tr(
            language,
            "pdf_period",
            start=_escape_reportlab(str(metrics["period_start"])),
            end=_escape_reportlab(str(metrics["period_end"])),
        )
    )

    metadata_line = (
        f"{tr(language, 'pdf_transactions_count', count=metrics['transactions_count'])}  |  "
        f"{tr(language, 'pdf_categories_count', count=metrics['categories_count'])}  |  "
        f"{tr(language, 'pdf_months_count', count=metrics['months_count'])}"
    )

    content = Paragraph(
        "<b>" + tr(language, "pdf_title") + "</b><br/>" + subtitle + "<br/><br/>" + metadata_line,
        styles["cover_subtitle"],
    )

    title = Paragraph(tr(language, "pdf_title"), styles["cover_title"])
    table = Table([[title], [content]], colWidths=[520])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1f3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 20),
                ("RIGHTPADDING", (0, 0), (-1, -1), 20),
                ("TOPPADDING", (0, 0), (-1, 0), 18),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 18),
            ]
        )
    )
    return table


def _build_premium_cover(
    metrics: dict[str, Any],
    styles: dict[str, ParagraphStyle],
    language: str,
) -> list[Any]:
    """Build a premium cover page with gradient-style banner and metadata badges."""
    elements: list[Any] = []
    elements.append(_build_cover_banner(metrics, styles, language))
    elements.append(Spacer(1, 16))
    elements.append(_build_kpi_dashboard(metrics, styles, language))
    return elements


def _build_kpi_table(metrics: dict[str, Any], styles: dict[str, ParagraphStyle], language: str) -> Table:
    cells = [
        (
            tr(language, "pdf_col_balance"),
            f"{float(metrics['balance']):.2f}",
            colors.HexColor("#d6f5e8"),
        ),
        (
            tr(language, "pdf_col_income"),
            f"{float(metrics['total_income']):.2f}",
            colors.HexColor("#dceeff"),
        ),
        (
            tr(language, "pdf_col_expense"),
            f"{float(metrics['total_expense']):.2f}",
            colors.HexColor("#ffe3de"),
        ),
        (
            tr(language, "pdf_transactions_count", count=metrics["transactions_count"]),
            str(metrics["transactions_count"]),
            colors.HexColor("#eef1ff"),
        ),
    ]

    data = [
        [
            Paragraph(f"<b>{cells[0][0]}</b><br/>{cells[0][1]}", styles["kpi_cell"]),
            Paragraph(f"<b>{cells[1][0]}</b><br/>{cells[1][1]}", styles["kpi_cell"]),
        ],
        [
            Paragraph(f"<b>{cells[2][0]}</b><br/>{cells[2][1]}", styles["kpi_cell"]),
            Paragraph(f"<b>{cells[3][0]}</b><br/>{cells[3][1]}", styles["kpi_cell"]),
        ],
    ]

    table = Table(data, colWidths=[260, 260])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, 0), cells[0][2]),
                ("BACKGROUND", (1, 0), (1, 0), cells[1][2]),
                ("BACKGROUND", (0, 1), (0, 1), cells[2][2]),
                ("BACKGROUND", (1, 1), (1, 1), cells[3][2]),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9fb1c9")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9fb1c9")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _build_kpi_dashboard(metrics: dict[str, Any], styles: dict[str, ParagraphStyle], language: str) -> Table:
    """Build an enhanced KPI dashboard with 6 cards in a 3x2 grid."""
    income = float(metrics["total_income"])
    expense = float(metrics["total_expense"])
    balance = float(metrics["balance"])
    savings_rate = _compute_savings_rate(income, expense)
    months_count = max(metrics["months_count"], 1)
    avg_expense = expense / months_count
    avg_income = income / months_count

    kpi_data = [
        (tr(language, "pdf_col_balance"), f"{balance:.2f}", "#d6f5e8" if balance >= 0 else "#ffe3de"),
        (tr(language, "pdf_col_income"), f"{income:.2f}", "#dceeff"),
        (tr(language, "pdf_col_expense"), f"{expense:.2f}", "#ffe3de"),
        (tr(language, "pdf_savings_rate", rate=savings_rate), f"{savings_rate:.1f}%", "#eef1ff"),
        (tr(language, "pdf_avg_expense_month", amount=avg_expense), f"{avg_expense:.2f}", "#fff3e0"),
        (tr(language, "pdf_avg_income_month", amount=avg_income), f"{avg_income:.2f}", "#e8f5e9"),
    ]

    rows = []
    for i in range(0, 6, 2):
        row = []
        for j in range(2):
            label_text, value_text, bg_color = kpi_data[i + j]
            cell_content = Paragraph(
                f"<b>{value_text}</b><br/><font size=8>{label_text}</font>",
                styles["kpi_cell"],
            )
            row.append(cell_content)
        rows.append(row)

    table = Table(rows, colWidths=[260, 260])
    style_cmds: list[tuple[Any, ...]] = [
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9fb1c9")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9fb1c9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]

    idx = 0
    for r in range(3):
        for c in range(2):
            style_cmds.append(("BACKGROUND", (c, r), (c, r), colors.HexColor(kpi_data[idx][2])))
            idx += 1

    table.setStyle(TableStyle(style_cmds))
    return table


def _top_category(category_rows: list[dict[str, Any]], key: str) -> str:
    if not category_rows:
        return "N/A"

    top_row = max(category_rows, key=lambda row: float(row[key]))
    top_value = float(top_row[key])
    if top_value <= 0:
        return "N/A"
    return f"{top_row['category']} ({top_value:.2f})"


def _chunks(items: list[list[str]], chunk_size: int) -> list[list[list[str]]]:
    return [items[index : index + chunk_size] for index in range(0, len(items), chunk_size)]


def _draw_page_number(canvas: Any, doc: Any, language: str) -> None:
    canvas.saveState()
    page_width = A4[0]

    # Header
    canvas.setStrokeColor(colors.HexColor("#1f3a5f"))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, A4[1] - 28, page_width - doc.rightMargin, A4[1] - 28)

    canvas.setFillColor(colors.HexColor("#1f3a5f"))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(doc.leftMargin, A4[1] - 24, tr(language, "pdf_footer_project"))
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(page_width / 2, A4[1] - 24, datetime.now().strftime("%d/%m/%Y"))
    canvas.drawRightString(page_width - doc.rightMargin, A4[1] - 24, tr(language, "pdf_footer_page", page=doc.page))

    # Footer
    canvas.setStrokeColor(colors.HexColor("#b8c4d6"))
    canvas.line(doc.leftMargin, 22, page_width - doc.rightMargin, 22)

    canvas.setFillColor(colors.HexColor("#556273"))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(doc.leftMargin, 10, tr(language, "pdf_footer_project"))
    canvas.drawRightString(page_width - doc.rightMargin, 10, tr(language, "pdf_footer_page", page=doc.page))
    canvas.restoreState()


# ---------------------------------------------------------------------------
# Helpers for dynamic aggregation (used by smart export & monthly reports)
# ---------------------------------------------------------------------------


def _compute_category_rows_from_transactions(transactions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories: dict[str, dict[str, float]] = defaultdict(lambda: {"income": 0.0, "expense": 0.0})
    for row in transactions:
        amount = float(row["amount"] or 0)
        cat = str(row["category"])
        if row["transaction_type"] == "income":
            categories[cat]["income"] += amount
        elif row["transaction_type"] == "expense":
            categories[cat]["expense"] += amount
    return [
        {
            "category": cat,
            "income": data["income"],
            "expense": data["expense"],
            "balance": data["income"] - data["expense"],
        }
        for cat, data in sorted(categories.items())
    ]


def _compute_month_rows_from_transactions(transactions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    months: dict[str, dict[str, float]] = defaultdict(lambda: {"income": 0.0, "expense": 0.0})
    for row in transactions:
        amount = float(row["amount"] or 0)
        month = str(row["transaction_date"])[:7]
        if row["transaction_type"] == "income":
            months[month]["income"] += amount
        elif row["transaction_type"] == "expense":
            months[month]["expense"] += amount
    return [
        {
            "month": month,
            "income": data["income"],
            "expense": data["expense"],
            "balance": data["income"] - data["expense"],
        }
        for month, data in sorted(months.items())
    ]


# ---------------------------------------------------------------------------
# JSON / YAML / HTML exporters
# ---------------------------------------------------------------------------


def _export_json(
    output_file: Path,
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    language: str,
) -> None:
    payload = {
        "meta": {
            "generated_at": datetime.now().isoformat(),
            "language": language,
            "record_count": len(transactions),
        },
        "transactions": transactions,
        "by_category": category_rows,
        "by_month": month_rows,
    }
    with output_file.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, default=str)


def _export_yaml(
    output_file: Path,
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    language: str,
) -> None:
    payload = {
        "meta": {
            "generated_at": datetime.now().isoformat(),
            "language": language,
            "record_count": len(transactions),
        },
        "transactions": transactions,
        "by_category": category_rows,
        "by_month": month_rows,
    }
    with output_file.open("w", encoding="utf-8") as handle:
        yaml.dump(payload, handle, allow_unicode=True, sort_keys=False, default_flow_style=False)


def _export_html(
    output_file: Path,
    transactions: list[dict[str, Any]],
    category_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    language: str,
) -> None:
    if not _PLOTLY_AVAILABLE:
        raise RuntimeError("Plotly is required for HTML export.")

    fig = make_subplots(
        rows=2,
        cols=1,
        subplot_titles=(
            tr(language, "pdf_table_category_title"),
            tr(language, "pdf_table_month_title"),
        ),
        vertical_spacing=0.12,
    )

    if category_rows:
        categories = [str(row["category"]) for row in category_rows]
        balances = [float(row["balance"] or 0) for row in category_rows]
        fig.add_trace(
            go.Bar(x=categories, y=balances, marker_color="steelblue", name=tr(language, "pdf_col_balance")),
            row=1,
            col=1,
        )

    if month_rows:
        months = [str(row["month"]) for row in month_rows]
        incomes = [float(row["income"] or 0) for row in month_rows]
        expenses = [float(row["expense"] or 0) for row in month_rows]
        fig.add_trace(
            go.Scatter(x=months, y=incomes, mode="lines+markers", name=tr(language, "pdf_col_income")),
            row=2,
            col=1,
        )
        fig.add_trace(
            go.Scatter(x=months, y=expenses, mode="lines+markers", name=tr(language, "pdf_col_expense")),
            row=2,
            col=1,
        )

    fig.update_layout(
        title_text=tr(language, "pdf_title"),
        height=800,
        template="plotly_white",
        showlegend=True,
    )

    fig.write_html(str(output_file), include_plotlyjs="cdn")


# ---------------------------------------------------------------------------
# Monthly consolidated PDF
# ---------------------------------------------------------------------------


def _export_monthly_pdf(
    output_file: Path,
    transactions: list[dict[str, Any]],
    language: str,
    year_month: str | None,
    budget_rows: list[dict[str, Any]] | None = None,
) -> None:
    from expenses_tracker.charts import generate_all_figures

    if not transactions:
        raise ValueError(tr(language, "warning_no_transactions_export"))

    target_month = year_month
    if target_month is None:
        target_month = max(str(row["transaction_date"])[:7] for row in transactions)

    month_transactions = [row for row in transactions if str(row["transaction_date"]).startswith(target_month)]
    if not month_transactions:
        raise ValueError(tr(language, "warning_no_transactions_export"))

    month_category_rows = _compute_category_rows_from_transactions(month_transactions)
    month_rows_local = _compute_month_rows_from_transactions(month_transactions)

    base_styles = getSampleStyleSheet()
    styles = _build_pdf_styles(base_styles)
    document = SimpleDocTemplate(
        str(output_file),
        pagesize=A4,
        leftMargin=28,
        rightMargin=28,
        topMargin=36,
        bottomMargin=26,
    )
    elements: list[Any] = []

    summary_metrics = _compute_summary_metrics(month_transactions, month_category_rows, month_rows_local)

    # Premium cover
    elements.append(_build_monthly_cover_banner(target_month, styles, language))
    elements.append(Spacer(1, 16))
    elements.append(_build_kpi_dashboard(summary_metrics, styles, language))
    elements.append(PageBreak())

    # Trend analysis
    elements.extend(_build_trend_analysis_section(summary_metrics, month_rows_local, styles, language))

    # Charts
    figures = generate_all_figures(month_category_rows, month_rows_local, language, "default")
    chart_title_map = {
        "line": "pdf_chart_evolution",
        "bar": "pdf_chart_categories",
        "pie": "pdf_chart_distribution",
        "scatter": "pdf_chart_scatter",
        "bar3d": "pdf_chart_3d",
        "forecast": "pdf_chart_forecast",
        "sankey": "pdf_chart_sankey",
    }
    for chart_key in ["line", "bar", "pie", "scatter", "bar3d", "forecast", "sankey"]:
        if chart_key in figures:
            elements.extend(_build_chart_section(
                figures[chart_key],
                chart_title_map[chart_key],
                styles,
                language,
            ))

    # Budget section for this month
    if budget_rows:
        elements.extend(_build_budget_section(budget_rows, styles, language))

    # Category table
    category_table_rows = [
        [
            _escape_reportlab(str(row["category"])),
            f"{float(row['income'] or 0):.2f}",
            f"{float(row['expense'] or 0):.2f}",
            f"{float(row['balance'] or 0):.2f}",
        ]
        for row in month_category_rows
    ]
    _append_paginated_table(
        elements=elements,
        styles=styles,
        title=tr(language, "pdf_table_category_title"),
        header=[
            tr(language, "excel_col_category"),
            tr(language, "pdf_col_income"),
            tr(language, "pdf_col_expense"),
            tr(language, "pdf_col_balance"),
        ],
        rows=category_table_rows,
        chunk_size=28,
        col_widths=[180, 95, 95, 95],
        text_column_indexes=[0],
        language=language,
    )

    # Transactions table
    movement_table_rows = [
        [
            _escape_reportlab(str(row["transaction_date"])),
            _escape_reportlab(_localize_transaction_type(str(row["transaction_type"]), language)),
            _escape_reportlab(str(row["category"])),
            f"{float(row['amount'] or 0):.2f}",
            _escape_reportlab(str(row["description"])),
        ]
        for row in month_transactions
    ]
    _append_paginated_table(
        elements=elements,
        styles=styles,
        title=tr(language, "pdf_table_transactions_title"),
        header=[
            tr(language, "excel_col_date"),
            tr(language, "excel_col_type"),
            tr(language, "excel_col_category"),
            tr(language, "excel_col_amount"),
            tr(language, "excel_col_description"),
        ],
        rows=movement_table_rows,
        chunk_size=22,
        col_widths=[86, 76, 102, 70, 136],
        text_column_indexes=[0, 1, 2, 4],
        language=language,
    )

    document.build(
        elements,
        onFirstPage=lambda canvas, doc: _draw_page_number(canvas, doc, language),
        onLaterPages=lambda canvas, doc: _draw_page_number(canvas, doc, language),
    )


def _build_monthly_cover_banner(
    year_month: str, styles: dict[str, ParagraphStyle], language: str
) -> Table:
    subtitle = (
        tr(language, "pdf_generated", timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        + "<br/>"
        + tr(language, "pdf_period", start=year_month, end=year_month)
    )
    content = Paragraph(
        "<b>" + tr(language, "monthly_report_title") + "</b><br/>" + subtitle,
        styles["cover_subtitle"],
    )

    title = Paragraph(tr(language, "monthly_report_title"), styles["cover_title"])
    table = Table([[title], [content]], colWidths=[520])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1f3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table
