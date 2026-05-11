"""Import transactions from CSV, Excel and JSON files.

Users can import data from external sources, preview the parsed rows
and confirm before inserting into the database.
"""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path
from typing import Any

from expenses_tracker.schemas import VALID_TRANSACTION_TYPES


def _safe_float(value: Any) -> float | None:
    """Parse a value to float, returning None on failure."""
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _safe_date(value: Any) -> date | None:
    """Parse a value to date, returning None on failure."""
    raw = str(value).strip()
    if not raw:
        return None
    for _fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            dt = raw.split("T")[0].split(" ")[0]
            return date.fromisoformat(dt)
        except ValueError:
            continue
    return None


def _col_val(key: str, headers: dict[str, int], row_values: tuple) -> str:
    """Extract a cell value by column key from a parsed Excel row."""
    idx = headers.get(key)
    if idx is None or idx >= len(row_values):
        return ""
    return str(row_values[idx] or "").strip()


def parse_csv(filepath: Path) -> list[dict[str, Any]]:
    """Parse a CSV file and return transaction-like dicts."""
    rows: list[dict[str, Any]] = []
    with filepath.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            amount = _safe_float(row.get("amount", row.get("Amount", 0)))
            if amount is None or amount <= 0:
                continue
            tx_type = str(row.get("transaction_type", row.get("type", "expense"))).strip().lower()
            if tx_type not in VALID_TRANSACTION_TYPES:
                tx_type = "expense"
            tx_date = _safe_date(row.get("transaction_date", row.get("date", row.get("Date", ""))))
            if tx_date is None:
                continue
            rows.append(
                {
                    "amount": amount,
                    "transaction_type": tx_type,
                    "category": str(row.get("category", row.get("Category", ""))).strip() or "other",
                    "transaction_date": tx_date.isoformat(),
                    "description": str(row.get("description", row.get("Description", ""))).strip(),
                    "currency": str(row.get("currency", row.get("Currency", "USD"))).strip().upper()[:3] or "USD",
                    "tags": str(row.get("tags", row.get("Tags", ""))).strip() or None,
                    "recurring": str(row.get("recurring", "false")).strip().lower() == "true",
                }
            )
    return rows


def parse_excel(filepath: Path) -> list[dict[str, Any]]:
    """Parse an Excel file and return transaction-like dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    headers: dict[str, int] = {}

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None:
            continue
        if r_idx == 1:
            for c_idx, cell in enumerate(row):
                key = str(cell or "").strip().lower()
                if key in {
                    "amount",
                    "type",
                    "transaction_type",
                    "category",
                    "date",
                    "transaction_date",
                    "description",
                    "currency",
                    "tags",
                    "recurring",
                }:
                    headers[key] = c_idx
            continue
        if not headers:
            continue

        row_values: tuple = row
        amount = _safe_float(_col_val("amount", headers, row_values))
        if amount is None or amount <= 0:
            continue
        tx_type = (
            _col_val("type", headers, row_values) or _col_val("transaction_type", headers, row_values) or "expense"
        )
        tx_type = tx_type.strip().lower()
        if tx_type not in VALID_TRANSACTION_TYPES:
            tx_type = "expense"
        tx_date = _safe_date(_col_val("date", headers, row_values) or _col_val("transaction_date", headers, row_values))
        if tx_date is None:
            continue
        rows.append(
            {
                "amount": amount,
                "transaction_type": tx_type,
                "category": _col_val("category", headers, row_values) or "other",
                "transaction_date": tx_date.isoformat(),
                "description": _col_val("description", headers, row_values),
                "currency": _col_val("currency", headers, row_values).upper()[:3] or "USD",
                "tags": _col_val("tags", headers, row_values) or None,
                "recurring": _col_val("recurring", headers, row_values).lower() == "true",
            }
        )
    wb.close()
    return rows


def parse_json(filepath: Path) -> list[dict[str, Any]]:
    """Parse a JSON file and return transaction-like dicts."""
    with filepath.open(encoding="utf-8") as f:
        data = json.load(f)

    items = data if isinstance(data, list) else data.get("transactions", data.get("data", []))
    rows: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        amount = _safe_float(item.get("amount", 0))
        if amount is None or amount <= 0:
            continue
        tx_type = str(item.get("transaction_type", item.get("type", "expense"))).strip().lower()
        if tx_type not in VALID_TRANSACTION_TYPES:
            tx_type = "expense"
        tx_date = _safe_date(item.get("transaction_date", item.get("date", "")))
        if tx_date is None:
            continue
        rows.append(
            {
                "amount": amount,
                "transaction_type": tx_type,
                "category": str(item.get("category", "")).strip() or "other",
                "transaction_date": tx_date.isoformat(),
                "description": str(item.get("description", "")).strip(),
                "currency": str(item.get("currency", "USD")).strip().upper()[:3] or "USD",
                "tags": str(item.get("tags", "")).strip() or None,
                "recurring": bool(item.get("recurring", False)),
            }
        )
    return rows


def import_transactions(filepath: str | Path) -> list[dict[str, Any]]:
    """Auto-detect format and parse transactions from a file."""
    path = Path(filepath)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_csv(path)
    elif suffix in {".xlsx", ".xls"}:
        return parse_excel(path)
    elif suffix == ".json":
        return parse_json(path)
    raise ValueError(f"Unsupported file format: {suffix}")


def import_and_save(filepath: str | Path, database: Any, language: str = "en") -> int:
    """Import from file, save to DB, return count of imported rows."""
    from expenses_tracker.schemas import TransactionInput

    parsed = import_transactions(filepath)
    count = 0
    for row in parsed:
        try:
            tx_date = date.fromisoformat(row.pop("transaction_date"))
            tx_input = TransactionInput(
                amount=float(row["amount"]),
                transaction_type=str(row["transaction_type"]),
                category=str(row["category"]),
                transaction_date=tx_date,
                description=str(row.get("description", "")),
                currency=str(row.get("currency", "USD")),
                tags=row.get("tags"),
                recurring=bool(row.get("recurring", False)),
            )
            database.add_transaction(tx_input, language=language)
            count += 1
        except (ValueError, KeyError):
            continue
    return count
